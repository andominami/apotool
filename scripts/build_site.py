#!/usr/bin/env python3
"""
保険算定ルール検索サイトのビルドスクリプト。

使い方:
    python3 scripts/build_site.py <入力エクセル.xlsx>

エクセルの「保険算定ルール一覧」シート(列: No, 日付, カテゴリ, タイトル, 内容(ルール詳細), 添付画像)
を読み込み、
    data/insurance_rules.json  … 抽出したデータ
    site/index.html            … 検索サイト本体(site/template.html にデータを埋め込んだもの)
を生成する。
"""
import datetime
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です。`pip install openpyxl` を実行してください。")

ROOT = Path(__file__).resolve().parent.parent
SHEET_NAME = "保険算定ルール一覧"


def load_rules(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"シート「{SHEET_NAME}」が見つかりません。シート一覧: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        no, date, category, title, content, _img = row
        if no is None and title is None:
            continue
        if isinstance(date, datetime.datetime):
            date = date.strftime("%Y/%m/%d")
        rules.append(
            {
                "no": no,
                "date": date,
                "category": category,
                "title": title,
                "content": content,
            }
        )
    return rules


def main():
    if len(sys.argv) != 2:
        sys.exit("使い方: python3 scripts/build_site.py <入力エクセル.xlsx>")

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        sys.exit(f"ファイルが見つかりません: {xlsx_path}")

    rules = load_rules(xlsx_path)
    print(f"{len(rules)} 件のルールを読み込みました。")

    data_path = ROOT / "data" / "insurance_rules.json"
    data_path.write_text(
        json.dumps(rules, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"書き出し: {data_path}")

    template_path = ROOT / "site" / "template.html"
    template = template_path.read_text(encoding="utf-8")
    html = template.replace("__DATA_JSON__", json.dumps(rules, ensure_ascii=False))

    out_path = ROOT / "site" / "index.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"書き出し: {out_path}")


if __name__ == "__main__":
    main()
